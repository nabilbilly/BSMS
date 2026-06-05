import time

from typing import List, Optional
import io
import csv
import openpyxl
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text, or_
from core.database import get_db
from core.config import settings
from models import models as db_models
from schemas import schemas as api_schemas
from core.sms_utils import normalize_ghana_phone, is_valid_phone
from worker.tasks import send_sms_task
from datetime import datetime

router = APIRouter()


@router.get("/health")
def get_infrastructure_health(db: Session = Depends(get_db)):
    """
    Real-time health check for all infrastructure services.
    Returns status of Redis, Celery Workers, and Database.
    """
    health = {}

    # 1. Redis Check
    try:
        import redis as redis_lib
        r = redis_lib.from_url(settings.REDIS_URL, socket_connect_timeout=3)
        start = time.time()
        r.ping()
        latency = round((time.time() - start) * 1000, 1)
        health["redis"] = {"status": "connected", "latency_ms": latency}
    except Exception as e:
        health["redis"] = {"status": "disconnected", "error": str(e)}

    # 2. Celery Worker Check
    try:
        from worker.tasks import celery_app
        inspector = celery_app.control.inspect(timeout=3)
        ping_result = inspector.ping()
        if ping_result:
            worker_names = list(ping_result.keys())
            health["celery_worker"] = {
                "status": "connected",
                "active_workers": len(worker_names),
                "workers": worker_names,
            }
        else:
            health["celery_worker"] = {
                "status": "disconnected",
                "error": "No workers responded to ping",
                "active_workers": 0,
            }
    except Exception as e:
        health["celery_worker"] = {
            "status": "disconnected",
            "error": str(e),
            "active_workers": 0,
        }

    # 3. Database Check
    try:
        start = time.time()
        db.execute(text("SELECT 1"))
        latency = round((time.time() - start) * 1000, 1)
        health["database"] = {"status": "connected", "latency_ms": latency}
    except Exception as e:
        health["database"] = {"status": "disconnected", "error": str(e)}

    # 4. Overall Summary
    services_down = sum(1 for s in health.values() if s["status"] == "disconnected")
    if services_down == 0:
        health["overall"] = "All Systems Operational"
    elif services_down == len(health):
        health["overall"] = "Critical: All Services Down"
    else:
        health["overall"] = f"{services_down} Service{'s' if services_down > 1 else ''} Down"

    return health


@router.post("/companies", response_model=api_schemas.Company)
def create_company(company_in: api_schemas.CompanyCreate, db: Session = Depends(get_db)):
    # Check if slug exists
    db_company = db.query(db_models.Company).filter(db_models.Company.slug == company_in.slug).first()
    if db_company:
        raise HTTPException(status_code=400, detail="Company slug already exists")
    
    new_company = db_models.Company(
        name=company_in.name,
        slug=company_in.slug,
        is_active=company_in.is_active,
        hubtel_client_id=company_in.hubtel_client_id,
        hubtel_client_secret=company_in.hubtel_client_secret,
        hubtel_sender_id=company_in.hubtel_sender_id,
        brevo_api_key=company_in.brevo_api_key,
        brevo_sender_email=company_in.brevo_sender_email,
        brevo_sender_name=company_in.brevo_sender_name
    )
    db.add(new_company)
    db.commit()
    db.refresh(new_company)
    return new_company

@router.get("/companies", response_model=List[api_schemas.Company])
def get_companies(db: Session = Depends(get_db)):
    return db.query(db_models.Company).all()

@router.get("/companies/{company_id}", response_model=api_schemas.Company)
def get_company(company_id: int, db: Session = Depends(get_db)):
    company = db.query(db_models.Company).filter(db_models.Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    return company

@router.put("/companies/{company_id}", response_model=api_schemas.Company)
def update_company(company_id: int, company_in: api_schemas.CompanyUpdate, db: Session = Depends(get_db)):
    company = db.query(db_models.Company).filter(db_models.Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    update_data = company_in.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(company, field, value)
    
    db.add(company)
    db.commit()
    db.refresh(company)
    return company
 
@router.delete("/companies/{company_id}")
def delete_company(company_id: int, db: Session = Depends(get_db)):
    company = db.query(db_models.Company).filter(db_models.Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    db.delete(company)
    db.commit()
    return {"detail": "Company and all associated data cleared"}

from core.security import get_pin_hash

@router.get("/stats", response_model=api_schemas.GlobalStats)
def get_global_stats(db: Session = Depends(get_db)):
    total_companies = db.query(db_models.Company).count()
    total_branches = db.query(db_models.Branch).count()
    total_sms = db.query(db_models.SMSLog).count()
    total_emails = db.query(db_models.EmailLog).count()
    total_customers = db.query(db_models.Customer).count()
    
    return {
        "total_companies": total_companies,
        "total_branches": total_branches,
        "total_sms": total_sms,
        "total_emails": total_emails,
        "total_customers": total_customers
    }

@router.get("/logs/sms", response_model=api_schemas.PaginatedAdminSMSLog)
def get_global_sms_logs(skip: int = 0, limit: int = 50, db: Session = Depends(get_db)):
    query = (
        db.query(db_models.SMSLog, db_models.Company.name, db_models.Branch.name)
        .join(db_models.Branch, db_models.SMSLog.branch_id == db_models.Branch.id)
        .join(db_models.Company, db_models.Branch.company_id == db_models.Company.id)
    )
    
    total = query.count()
    results = query.order_by(db_models.SMSLog.created_at.desc()).offset(skip).limit(limit).all()
    
    items = []
    for log, company_name, branch_name in results:
        log_dict = {c.name: getattr(log, c.name) for c in log.__table__.columns}
        log_dict["company_name"] = company_name
        log_dict["branch_name"] = branch_name
        items.append(log_dict)
        
    return {"items": items, "total": total}

@router.get("/logs/email", response_model=api_schemas.PaginatedAdminEmailLog)
def get_global_email_logs(skip: int = 0, limit: int = 50, db: Session = Depends(get_db)):
    query = (
        db.query(db_models.EmailLog, db_models.Company.name, db_models.Branch.name)
        .join(db_models.Branch, db_models.EmailLog.branch_id == db_models.Branch.id)
        .join(db_models.Company, db_models.Branch.company_id == db_models.Company.id)
    )
    
    total = query.count()
    results = query.order_by(db_models.EmailLog.created_at.desc()).offset(skip).limit(limit).all()
    
    items = []
    for log, company_name, branch_name in results:
        log_dict = {c.name: getattr(log, c.name) for c in log.__table__.columns}
        log_dict["company_name"] = company_name
        log_dict["branch_name"] = branch_name
        items.append(log_dict)
        
    return {"items": items, "total": total}

@router.get("/contacts", response_model=api_schemas.PaginatedAdminCustomer)
def get_global_contacts(
    skip: int = 0,
    limit: int = 50,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = (
        db.query(db_models.Customer, db_models.Company.name, db_models.Branch.name)
        .join(db_models.Branch, db_models.Customer.branch_id == db_models.Branch.id)
        .join(db_models.Company, db_models.Branch.company_id == db_models.Company.id)
    )
    
    if search:
        query = query.filter(
            or_(
                db_models.Customer.full_name.ilike(f"%{search}%"),
                db_models.Customer.phone_number.ilike(f"%{search}%"),
                db_models.Customer.email.ilike(f"%{search}%"),
                db_models.Branch.name.ilike(f"%{search}%"),
                db_models.Company.name.ilike(f"%{search}%")
            )
        )
        
    total = query.count()
    results = query.order_by(db_models.Customer.created_at.desc()).offset(skip).limit(limit).all()
    
    items = []
    for customer, company_name, branch_name in results:
        items.append({
            "id": customer.id,
            "full_name": customer.full_name,
            "phone_number": customer.phone_number,
            "email": customer.email,
            "branch_name": branch_name,
            "company_name": company_name,
            "created_at": customer.created_at
        })
        
    return {"items": items, "total": total}

@router.get("/contacts/export")
def export_global_contacts(
    format: str = "csv",
    db: Session = Depends(get_db)
):
    query = (
        db.query(db_models.Customer, db_models.Company.name, db_models.Branch.name)
        .join(db_models.Branch, db_models.Customer.branch_id == db_models.Branch.id)
        .join(db_models.Company, db_models.Branch.company_id == db_models.Company.id)
        .order_by(db_models.Customer.created_at.desc())
    )
    
    results = query.all()
    
    if format.lower() == "xlsx":
        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contacts"
        
        # Headers
        headers = ["Name", "Phone Number", "Email", "Company", "Branch", "Signup Date"]
        ws.append(headers)
        
        for customer, company_name, branch_name in results:
            ws.append([
                customer.full_name,
                customer.phone_number,
                customer.email or "",
                company_name,
                branch_name,
                customer.created_at.strftime("%Y-%m-%d %H:%M:%S") if customer.created_at else ""
            ])
            
        # Write to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers = {
            'Content-Disposition': 'attachment; filename="contacts.xlsx"'
        }
        return StreamingResponse(
            output,
            headers=headers,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        # Default to CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow(["Name", "Phone Number", "Email", "Company", "Branch", "Signup Date"])
        
        for customer, company_name, branch_name in results:
            writer.writerow([
                customer.full_name,
                customer.phone_number,
                customer.email or "",
                company_name,
                branch_name,
                customer.created_at.strftime("%Y-%m-%d %H:%M:%S") if customer.created_at else ""
            ])
            
        output.seek(0)
        
        headers = {
            'Content-Disposition': 'attachment; filename="contacts.csv"'
        }
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8")),
            headers=headers,
            media_type="text/csv"
        )

@router.post("/contacts/import")
async def import_global_contacts(
    file: UploadFile = File(...),
    default_company_id: Optional[int] = Form(None),
    default_branch_id: Optional[int] = Form(None),
    db: Session = Depends(get_db)
):
    try:
        contents = await file.read()
        rows_data = [] # List of dicts
        
        filename = file.filename.lower()
        if filename.endswith(".csv"):
            stream = io.StringIO(contents.decode("utf-8"))
            reader = csv.reader(stream)
            rows = list(reader)
            if not rows:
                raise HTTPException(status_code=400, detail="Empty CSV file")
            
            # Find header indices
            headers = [h.strip().lower() for h in rows[0]]
            
            # Map headers to indices
            name_idx = next((i for i, h in enumerate(headers) if h in ["name", "full_name", "full name"]), -1)
            phone_idx = next((i for i, h in enumerate(headers) if h in ["phone", "phone_number", "phone number"]), -1)
            email_idx = next((i for i, h in enumerate(headers) if h in ["email", "email_address", "email address"]), -1)
            company_idx = next((i for i, h in enumerate(headers) if h in ["company", "company_slug", "company_name"]), -1)
            branch_idx = next((i for i, h in enumerate(headers) if h in ["branch", "branch_code", "branch_name"]), -1)
            
            # If name or phone headers aren't explicitly matched, assume column 0 is phone/name
            if name_idx == -1 and len(headers) > 1:
                name_idx = 1
            if phone_idx == -1:
                phone_idx = 0
                
            start_row = 1 if len(headers) > 0 and (name_idx != -1 or phone_idx != -1) else 0
            
            for row in rows[start_row:]:
                if not row:
                    continue
                name = row[name_idx].strip() if name_idx != -1 and name_idx < len(row) else ""
                phone = row[phone_idx].strip() if phone_idx != -1 and phone_idx < len(row) else ""
                email = row[email_idx].strip() if email_idx != -1 and email_idx < len(row) else None
                company = row[company_idx].strip() if company_idx != -1 and company_idx < len(row) else None
                branch = row[branch_idx].strip() if branch_idx != -1 and branch_idx < len(row) else None
                
                if phone:
                    rows_data.append({
                        "name": name or "Customer",
                        "phone": phone,
                        "email": email,
                        "company": company,
                        "branch": branch
                    })
                    
        elif filename.endswith((".xlsx", ".xls")):
            wb = openpyxl.load_workbook(io.BytesIO(contents))
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise HTTPException(status_code=400, detail="Empty Excel file")
                
            # Header matching
            headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
            name_idx = next((i for i, h in enumerate(headers) if h in ["name", "full_name", "full name"]), -1)
            phone_idx = next((i for i, h in enumerate(headers) if h in ["phone", "phone_number", "phone number"]), -1)
            email_idx = next((i for i, h in enumerate(headers) if h in ["email", "email_address", "email address"]), -1)
            company_idx = next((i for i, h in enumerate(headers) if h in ["company", "company_slug", "company_name"]), -1)
            branch_idx = next((i for i, h in enumerate(headers) if h in ["branch", "branch_code", "branch_name"]), -1)
            
            if name_idx == -1 and len(headers) > 1:
                name_idx = 1
            if phone_idx == -1:
                phone_idx = 0
                
            start_row = 1
            for row in rows[start_row:]:
                if not row or row[0] is None:
                    continue
                name = str(row[name_idx]).strip() if name_idx != -1 and name_idx < len(row) and row[name_idx] is not None else ""
                phone = str(row[phone_idx]).strip() if phone_idx != -1 and phone_idx < len(row) and row[phone_idx] is not None else ""
                email = str(row[email_idx]).strip() if email_idx != -1 and email_idx < len(row) and row[email_idx] is not None else None
                company = str(row[company_idx]).strip() if company_idx != -1 and company_idx < len(row) and row[company_idx] is not None else None
                branch = str(row[branch_idx]).strip() if branch_idx != -1 and branch_idx < len(row) and row[branch_idx] is not None else None
                
                if phone:
                    rows_data.append({
                        "name": name or "Customer",
                        "phone": phone,
                        "email": email,
                        "company": company,
                        "branch": branch
                    })
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Please use CSV or Excel.")
            
        if not rows_data:
            raise HTTPException(status_code=400, detail="No valid contact records found in file")
            
        imported_count = 0
        skipped_count = 0
        
        # Load all companies and branches cache to speed up lookup
        companies_cache = {c.slug.lower(): c.id for c in db.query(db_models.Company).all()}
        companies_name_cache = {c.name.lower(): c.id for c in db.query(db_models.Company).all()}
        
        # Cache branches
        branches_list = db.query(db_models.Branch).all()
        branches_code_cache = {(b.company_id, b.branch_code.lower()): b.id for b in branches_list}
        branches_name_cache = {(b.company_id, b.name.lower()): b.id for b in branches_list}
        
        for r_data in rows_data:
            norm_phone = normalize_ghana_phone(r_data["phone"])
            if not is_valid_phone(norm_phone):
                skipped_count += 1
                continue
                
            # Determine target company and branch
            company_id = None
            if r_data["company"]:
                c_str = r_data["company"].lower()
                company_id = companies_cache.get(c_str) or companies_name_cache.get(c_str)
            
            if not company_id:
                company_id = default_company_id
                
            if not company_id:
                skipped_count += 1
                continue
                
            branch_id = None
            if r_data["branch"]:
                b_str = r_data["branch"].lower()
                branch_id = branches_code_cache.get((company_id, b_str)) or branches_name_cache.get((company_id, b_str))
                
            if not branch_id:
                branch_id = default_branch_id
                
            if not branch_id:
                # Try to pick first branch of company
                fb = db.query(db_models.Branch).filter(db_models.Branch.company_id == company_id).first()
                if fb:
                    branch_id = fb.id
                    
            if not branch_id:
                skipped_count += 1
                continue
                
            # Check if customer already exists for this branch with normalized phone
            existing = (
                db.query(db_models.Customer)
                .filter(
                    db_models.Customer.phone_number == norm_phone,
                    db_models.Customer.branch_id == branch_id
                )
                .first()
            )
            
            if existing:
                existing.full_name = r_data["name"]
                if r_data["email"]:
                    existing.email = r_data["email"]
                imported_count += 1
            else:
                new_cust = db_models.Customer(
                    full_name=r_data["name"],
                    phone_number=norm_phone,
                    email=r_data["email"],
                    branch_id=branch_id
                )
                db.add(new_cust)
                imported_count += 1
                
        db.commit()
        return {"status": "success", "imported_count": imported_count, "skipped_count": skipped_count}
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to process imports: {str(e)}")

@router.post("/companies/{company_id}/branches", response_model=api_schemas.Branch)
def create_company_branch(company_id: int, branch_in: api_schemas.BranchCreate, db: Session = Depends(get_db)):
    # Check if company exists
    company = db.query(db_models.Company).filter(db_models.Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
        
    # Check if branch code exists
    existing = db.query(db_models.Branch).filter(db_models.Branch.branch_code == branch_in.branch_code).first()
    if existing:
        raise HTTPException(status_code=400, detail="Branch code already exists")
        
    new_branch = db_models.Branch(
        name=branch_in.name,
        branch_code=branch_in.branch_code,
        pin=get_pin_hash(branch_in.pin),
        company_id=company_id
    )
    db.add(new_branch)
    db.commit()
    db.refresh(new_branch)
    return new_branch

@router.get("/public/companies/{slug}/branches", response_model=List[api_schemas.Branch])
def get_company_branches(slug: str, db: Session = Depends(get_db)):
    company = db.query(db_models.Company).filter(db_models.Company.slug == slug).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    return company.branches

@router.post("/bulk-sms-import")
async def import_and_send_bulk_sms(
    company_id: int = Form(...),
    branch_id: int = Form(...),
    message_content: str = Form(...),
    scheduled_for: Optional[datetime] = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    try:
        # 1. Validate Company and Branch
        branch = db.query(db_models.Branch).filter(
            db_models.Branch.id == branch_id,
            db_models.Branch.company_id == company_id
        ).first()
        
        if not branch:
            raise HTTPException(status_code=404, detail="Branch not found or doesn't belong to the company")

        # 2. Parse File
        contents = await file.read()
        recipients = [] # List of dicts: {"phone": ..., "name": ...}

        filename = file.filename.lower()
        if filename.endswith(".csv"):
            stream = io.StringIO(contents.decode("utf-8"))
            reader = csv.reader(stream)
            # Try to detect header
            rows = list(reader)
            if not rows:
                 raise HTTPException(status_code=400, detail="Empty CSV file")
            
            # Simple heuristic: if first row looks like data, it's not a header
            # We'll assume columns are Phone, Name
            start_idx = 0
            if "phone" in rows[0][0].lower() or "name" in rows[0][0].lower():
                start_idx = 1
                
            for row in rows[start_idx:]:
                if len(row) >= 1:
                    phone = row[0].strip()
                    name = row[1].strip() if len(row) >= 2 else ""
                    recipients.append({"phone": phone, "name": name})

        elif filename.endswith((".xlsx", ".xls")):
            wb = openpyxl.load_workbook(io.BytesIO(contents))
            sheet = wb.active
            # Assume Col A: Phone, Col B: Name
            for row in sheet.iter_rows(min_row=1, values_only=True):
                if not row[0]: continue
                # Skip header if any
                if str(row[0]).lower() in ["phone", "recipient", "number"]:
                    continue
                
                recipients.append({
                    "phone": str(row[0]).strip(),
                    "name": str(row[1]).strip() if len(row) >= 2 and row[1] else ""
                })
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Please use CSV or Excel.")

        if not recipients:
            raise HTTPException(status_code=400, detail="No valid recipients found in file")

        # 3. Process Recipients
        queued_count = 0
        tasks_to_queue = []
        seen_phones = set()
        for rec in recipients:
            norm_phone = normalize_ghana_phone(rec["phone"])
            if not is_valid_phone(norm_phone):
                continue # Skip invalid numbers
            
            if norm_phone in seen_phones:
                continue # Skip duplicates
            seen_phones.add(norm_phone)
            
            # Placeholder replacement
            final_content = message_content.replace("{name}", rec["name"] or "Customer")
            final_content = final_content.replace("{branch}", branch.name)
            
            # Create Log
            new_log = db_models.SMSLog(
                branch_id=branch.id,
                phone_number=norm_phone,
                message_type="bulk",
                message_content=final_content,
                status="queued",
                is_bulk=True,
                scheduled_for=scheduled_for or datetime.utcnow()
            )
            db.add(new_log)
            db.flush() # Get ID

            tasks_to_queue.append((norm_phone, final_content, new_log.id))
            queued_count += 1
        
        db.commit()

        # Queue tasks after commit is successful
        for norm_phone, final_content, log_id in tasks_to_queue:
            if scheduled_for and scheduled_for > datetime.utcnow():
                 send_sms_task.apply_async(args=[norm_phone, final_content, log_id], eta=scheduled_for)
            else:
                 send_sms_task.delay(norm_phone, final_content, log_id)
        return {"status": "success", "queued_count": queued_count}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to process bulk import: {str(e)}")
